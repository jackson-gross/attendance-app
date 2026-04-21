from flask import Flask, request, jsonify, render_template, send_file, session, redirect, url_for
from flask_sqlalchemy import SQLAlchemy
from datetime import datetime
from functools import wraps
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import hashlib
import io
import os

app = Flask(__name__, static_folder='static', static_url_path='/static')
app.secret_key = os.environ.get('SECRET_KEY', 'change-me-in-production-please')
db_url = os.environ.get('DATABASE_URL', 'sqlite:///attendance.db')
app.config['SQLALCHEMY_DATABASE_URI'] = db_url
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
db = SQLAlchemy(app)


# ---------------------------------------------------------------------------
# Models
# ---------------------------------------------------------------------------

class Person(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False)
    email = db.Column(db.String(120), unique=True, nullable=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    records = db.relationship('AttendanceRecord', backref='person', lazy=True, cascade='all, delete-orphan')


class AttendanceRecord(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    person_id = db.Column(db.Integer, db.ForeignKey('person.id'), nullable=False)
    date = db.Column(db.Date, nullable=False)
    session = db.Column(db.String(2), nullable=False, default='AM')   # 'AM' or 'PM'
    status = db.Column(db.String(20), nullable=False, default='present')
    note = db.Column(db.String(255), nullable=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    __table_args__ = (db.UniqueConstraint('person_id', 'date', 'session', name='unique_person_date_session'),)


class AdminCredentials(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(80), nullable=False, unique=True)
    password_hash = db.Column(db.String(128), nullable=False)

    @staticmethod
    def hash_password(pw):
        return hashlib.sha256(pw.encode()).hexdigest()

    def check_password(self, pw):
        return self.password_hash == self.hash_password(pw)


def init_db():
    db.create_all()
    # Seed default admin if none exists
    if not AdminCredentials.query.first():
        admin = AdminCredentials(
            username='admin',
            password_hash=AdminCredentials.hash_password('admin')
        )
        db.session.add(admin)
        db.session.commit()


with app.app_context():
    init_db()


# ---------------------------------------------------------------------------
# Auth helpers
# ---------------------------------------------------------------------------

def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not session.get('logged_in'):
            if request.is_json:
                return jsonify({'error': 'Unauthorized'}), 401
            return redirect(url_for('login_page'))
        return f(*args, **kwargs)
    return decorated


# ---------------------------------------------------------------------------
# Auth routes
# ---------------------------------------------------------------------------

@app.route('/login', methods=['GET'])
def login_page():
    if session.get('logged_in'):
        return redirect(url_for('index'))
    return render_template('login.html')


@app.route('/api/auth/login', methods=['POST'])
def do_login():
    data = request.json
    admin = AdminCredentials.query.filter_by(username=data.get('username', '')).first()
    if admin and admin.check_password(data.get('password', '')):
        session['logged_in'] = True
        session['username'] = admin.username
        return jsonify({'success': True})
    return jsonify({'error': 'Invalid username or password'}), 401


@app.route('/api/auth/logout', methods=['POST'])
def do_logout():
    session.clear()
    return jsonify({'success': True})


@app.route('/api/auth/change-password', methods=['POST'])
@login_required
def change_password():
    data = request.json
    admin = AdminCredentials.query.filter_by(username=session['username']).first()
    if not admin or not admin.check_password(data.get('current_password', '')):
        return jsonify({'error': 'Current password is incorrect'}), 400
    new_pw = data.get('new_password', '')
    if len(new_pw) < 4:
        return jsonify({'error': 'New password must be at least 4 characters'}), 400
    admin.password_hash = AdminCredentials.hash_password(new_pw)
    db.session.commit()
    return jsonify({'success': True})


@app.route('/api/auth/change-username', methods=['POST'])
@login_required
def change_username():
    data = request.json
    new_username = data.get('new_username', '').strip()
    if not new_username:
        return jsonify({'error': 'Username cannot be empty'}), 400
    if AdminCredentials.query.filter_by(username=new_username).first():
        return jsonify({'error': 'Username already taken'}), 400
    admin = AdminCredentials.query.filter_by(username=session['username']).first()
    if not admin or not admin.check_password(data.get('password', '')):
        return jsonify({'error': 'Password is incorrect'}), 400
    admin.username = new_username
    db.session.commit()
    session['username'] = new_username
    return jsonify({'success': True})


@app.route('/api/auth/me', methods=['GET'])
@login_required
def auth_me():
    return jsonify({'username': session.get('username')})


# ---------------------------------------------------------------------------
# Main app
# ---------------------------------------------------------------------------

@app.route('/')
@login_required
def index():
    return render_template('index.html')


# ---------------------------------------------------------------------------
# People
# ---------------------------------------------------------------------------

@app.route('/api/people', methods=['GET'])
@login_required
def get_people():
    people = Person.query.order_by(Person.name).all()
    return jsonify([{'id': p.id, 'name': p.name, 'email': p.email, 'created_at': p.created_at.isoformat()} for p in people])


@app.route('/api/people', methods=['POST'])
@login_required
def add_person():
    data = request.json
    if not data.get('name'):
        return jsonify({'error': 'Name is required'}), 400
    if data.get('email') and Person.query.filter_by(email=data['email']).first():
        return jsonify({'error': 'Email already exists'}), 400
    p = Person(name=data['name'], email=data.get('email') or None)
    db.session.add(p)
    db.session.commit()
    return jsonify({'id': p.id, 'name': p.name, 'email': p.email}), 201


@app.route('/api/people/<int:pid>', methods=['PUT'])
@login_required
def update_person(pid):
    p = Person.query.get_or_404(pid)
    data = request.json
    if 'name' in data: p.name = data['name']
    if 'email' in data: p.email = data['email'] or None
    db.session.commit()
    return jsonify({'id': p.id, 'name': p.name, 'email': p.email})


@app.route('/api/people/<int:pid>', methods=['DELETE'])
@login_required
def delete_person(pid):
    p = Person.query.get_or_404(pid)
    db.session.delete(p)
    db.session.commit()
    return jsonify({'success': True})


# ---------------------------------------------------------------------------
# Attendance
# ---------------------------------------------------------------------------

@app.route('/api/attendance', methods=['GET'])
@login_required
def get_attendance():
    date_str = request.args.get('date')
    session_filter = request.args.get('session')
    person_id = request.args.get('person_id')
    q = AttendanceRecord.query.join(Person)
    if date_str:
        try:
            q = q.filter(AttendanceRecord.date == datetime.strptime(date_str, '%Y-%m-%d').date())
        except ValueError:
            return jsonify({'error': 'Invalid date'}), 400
    if session_filter:
        q = q.filter(AttendanceRecord.session == session_filter.upper())
    if person_id:
        q = q.filter(AttendanceRecord.person_id == person_id)
    q = q.order_by(AttendanceRecord.date.desc(), AttendanceRecord.session, Person.name)
    return jsonify([{
        'id': r.id,
        'person_id': r.person_id,
        'person_name': r.person.name,
        'date': r.date.isoformat(),
        'session': r.session,
        'status': r.status,
        'note': r.note
    } for r in q.all()])


@app.route('/api/attendance/bulk', methods=['POST'])
@login_required
def bulk_attendance():
    data = request.json
    date_str = data.get('date')
    sess = data.get('session', 'AM').upper()
    if not date_str:
        return jsonify({'error': 'date required'}), 400
    try:
        d = datetime.strptime(date_str, '%Y-%m-%d').date()
    except ValueError:
        return jsonify({'error': 'Invalid date'}), 400
    for item in data.get('records', []):
        r = AttendanceRecord.query.filter_by(person_id=item['person_id'], date=d, session=sess).first()
        if r:
            r.status = item['status']
            r.note = item.get('note', r.note)
        else:
            r = AttendanceRecord(
                person_id=item['person_id'], date=d, session=sess,
                status=item['status'], note=item.get('note')
            )
            db.session.add(r)
    db.session.commit()
    return jsonify({'success': True, 'count': len(data.get('records', []))})


@app.route('/api/attendance/<int:rid>', methods=['DELETE'])
@login_required
def delete_attendance(rid):
    r = AttendanceRecord.query.get_or_404(rid)
    db.session.delete(r)
    db.session.commit()
    return jsonify({'success': True})


@app.route('/api/attendance/summary', methods=['GET'])
@login_required
def attendance_summary():
    start_str = request.args.get('start')
    end_str = request.args.get('end')
    sess = request.args.get('session')  # optional filter
    people = Person.query.order_by(Person.name).all()
    summary = []
    for person in people:
        q = AttendanceRecord.query.filter_by(person_id=person.id)
        if start_str:
            q = q.filter(AttendanceRecord.date >= datetime.strptime(start_str, '%Y-%m-%d').date())
        if end_str:
            q = q.filter(AttendanceRecord.date <= datetime.strptime(end_str, '%Y-%m-%d').date())
        if sess:
            q = q.filter(AttendanceRecord.session == sess.upper())
        am_counts = {'present': 0, 'absent': 0}
        pm_counts = {'present': 0, 'absent': 0}
        for r in q.all():
            if r.session == 'AM':
                am_counts[r.status] = am_counts.get(r.status, 0) + 1
            else:
                pm_counts[r.status] = pm_counts.get(r.status, 0) + 1
        summary.append({
            'name': person.name,
            'am_present': am_counts['present'], 'am_absent': am_counts['absent'],
            'pm_present': pm_counts['present'], 'pm_absent': pm_counts['absent'],
        })
    return jsonify(summary)


# ---------------------------------------------------------------------------
# Export
# ---------------------------------------------------------------------------

@app.route('/api/export', methods=['GET'])
@login_required
def export_excel():
    start_str = request.args.get('start')
    end_str = request.args.get('end')
    date_str = request.args.get('date')

    wb = openpyxl.Workbook()
    hdr_fill = PatternFill('solid', start_color='1A1A2E')
    sub_fill = PatternFill('solid', start_color='16213E')
    alt_fill = PatternFill('solid', start_color='F8F9FA')
    wht_fill = PatternFill('solid', start_color='FFFFFF')
    am_fill = PatternFill('solid', start_color='EDE9FE')
    pm_fill = PatternFill('solid', start_color='DBEAFE')
    present_fill = PatternFill('solid', start_color='D4EDDA')
    absent_fill = PatternFill('solid', start_color='F8D7DA')
    border = Border(
        left=Side(style='thin', color='DDDDDD'), right=Side(style='thin', color='DDDDDD'),
        top=Side(style='thin', color='DDDDDD'), bottom=Side(style='thin', color='DDDDDD')
    )

    def apply_header(ws, title, cols):
        ws.merge_cells(f'A1:{get_column_letter(len(cols))}1')
        c = ws['A1']
        c.value = title
        c.font = Font(name='Arial', bold=True, size=14, color='FFFFFF')
        c.fill = hdr_fill
        c.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 34
        for i, h in enumerate(cols, 1):
            cell = ws.cell(row=2, column=i, value=h)
            cell.font = Font(name='Arial', bold=True, size=10, color='FFFFFF')
            cell.fill = sub_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        ws.row_dimensions[2].height = 20

    export_ts = datetime.now().strftime('%B %d, %Y at %I:%M %p')

    # Sheet 1: Log
    ws1 = wb.active
    ws1.title = 'Attendance Log'
    apply_header(ws1, f'Attendance Report — Exported {export_ts}',
                 ['Date', 'Session', 'Name', 'Status', 'Note', 'Recorded At'])

    q = AttendanceRecord.query.join(Person).order_by(
        AttendanceRecord.date.desc(), AttendanceRecord.session, Person.name)
    if date_str:
        q = q.filter(AttendanceRecord.date == datetime.strptime(date_str, '%Y-%m-%d').date())
    elif start_str and end_str:
        q = q.filter(AttendanceRecord.date.between(
            datetime.strptime(start_str, '%Y-%m-%d').date(),
            datetime.strptime(end_str, '%Y-%m-%d').date()))
    records = q.all()

    for i, r in enumerate(records, 3):
        rf = alt_fill if i % 2 == 0 else wht_fill
        sess_fill = am_fill if r.session == 'AM' else pm_fill
        status_fill = present_fill if r.status == 'present' else absent_fill
        vals = [r.date.strftime('%Y-%m-%d'), r.session, r.person.name,
                r.status.capitalize(), r.note or '', r.created_at.strftime('%Y-%m-%d %H:%M')]
        for col, val in enumerate(vals, 1):
            cell = ws1.cell(row=i, column=col, value=val)
            cell.font = Font(name='Arial', size=10, bold=(col == 4))
            cell.border = border
            cell.alignment = Alignment(vertical='center')
            if col == 2:
                cell.fill = sess_fill
            elif col == 4:
                cell.fill = status_fill
            else:
                cell.fill = rf

    for i, w in enumerate([14, 10, 28, 12, 30, 18], 1):
        ws1.column_dimensions[get_column_letter(i)].width = w

    # Sheet 2: Summary
    ws2 = wb.create_sheet('Summary')
    apply_header(ws2, f'Attendance Summary — Exported {export_ts}',
                 ['Name', 'AM Present', 'AM Absent', 'AM Rate%', 'PM Present', 'PM Absent', 'PM Rate%'])

    for i, person in enumerate(Person.query.order_by(Person.name).all(), 3):
        pq = AttendanceRecord.query.filter_by(person_id=person.id)
        if date_str:
            pq = pq.filter(AttendanceRecord.date == datetime.strptime(date_str, '%Y-%m-%d').date())
        elif start_str and end_str:
            pq = pq.filter(AttendanceRecord.date.between(
                datetime.strptime(start_str, '%Y-%m-%d').date(),
                datetime.strptime(end_str, '%Y-%m-%d').date()))
        am = {'present': 0, 'absent': 0}
        pm = {'present': 0, 'absent': 0}
        for rec in pq.all():
            if rec.session == 'AM':
                am[rec.status] = am.get(rec.status, 0) + 1
            else:
                pm[rec.status] = pm.get(rec.status, 0) + 1
        am_total = am['present'] + am['absent']
        pm_total = pm['present'] + pm['absent']
        am_rate = f"{round(am['present']/am_total*100,1)}%" if am_total else '—'
        pm_rate = f"{round(pm['present']/pm_total*100,1)}%" if pm_total else '—'
        rf = alt_fill if i % 2 == 0 else wht_fill
        for col, val in enumerate([person.name, am['present'], am['absent'], am_rate,
                                    pm['present'], pm['absent'], pm_rate], 1):
            cell = ws2.cell(row=i, column=col, value=val)
            cell.font = Font(name='Arial', size=10)
            cell.fill = rf
            cell.border = border
            cell.alignment = Alignment(vertical='center', horizontal='left' if col == 1 else 'center')

    for i, w in enumerate([28, 12, 12, 10, 12, 12, 10], 1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    ts = datetime.now().strftime('%Y%m%d_%H%M%S')
    return send_file(out, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=f'attendance_{ts}.xlsx')


if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000, debug=False)
